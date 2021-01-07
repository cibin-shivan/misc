import openpyxl
from openpyxl import Workbook
import re

def parse_sent_report():
    # Input file
    wb_input = openpyxl.load_workbook(filename=r'<Input file path>', read_only=False)
    ws_input = wb_input.worksheets[0]

    # Output file
    wb_output = Workbook()
    ws_output = wb_output.create_sheet(title='Data', index=0)

    # Declare variables
    question = ''
    tot_row_no = 0
    start = False
    first = True
    response = ''
    date = ''

    # Row iterate
    for r in range(1, ws_input.max_row):
        # print(str(ws.cell(r, 1).value))

        # Column iterate
        for c in range(1, ws_input.max_column):
            # Fetch question logic
            if re.search('QC1*', str(ws_input.cell(r, 1).value)):
                question = str(ws_input.cell(r, c).value)
                break
            # Identify Total row logic
            elif str(ws_input.cell(r, 1).value) == "Total":
                tot_row_no = r
                start = True
                break
            # Identify Total row logic
            elif str(ws_input.cell(r, 1).value) == "None":
                start = False
                break
            elif not start:
                break

            response = str(ws_input.cell(r, 1).value)

            if start and c > 1 and c % 2 == 0:
                # print(r, c, question, response, str(ws_input.cell(tot_row_no - 1, c).value),
                #       str(ws_input.cell(tot_row_no, c).value), str(ws_input.cell(r, c).value), str(ws_input.cell(r, c + 1).value))

                if first:
                    first = False
                    row = ('Question', 'Response', 'Date','Count','Percent','Comment')
                    ws_output.append(row)
                else:
                    row = (question, response, str(ws_input.cell(tot_row_no-1, c).value), str(ws_input.cell(tot_row_no, c).value), str(ws_input.cell(r, c).value), str(ws_input.cell(r, c+1).value))
                    ws_output.append(row)

        wb_output.save(filename="<Output file path>")


